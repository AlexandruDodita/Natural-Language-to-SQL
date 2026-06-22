import io
import os
import json
import logging
import httpx
import sqlparse
from sqlparse.sql import Statement
from sqlparse.tokens import Keyword, DDL, DML
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
import google.generativeai as genai
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, AreaChart, Reference
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "")
GEMINI_MODEL = os.getenv("GEMINI_MODEL", "gemini-2.0-flash")
SQL_BACKEND_URL = os.getenv("SQL_BACKEND_URL", "http://test_app_backend:8080")

app = FastAPI(title="RAG Service")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------------------------------------------------------------------
# Schema context — baked in at startup
# ---------------------------------------------------------------------------
SCHEMA = """
Database: car_rental (PostgreSQL)

Tables:

locations(id, name, city, address, phone, created_at)
  - Branch offices. PK: id.

employees(id, location_id FK→locations, first_name, last_name,
          role CHECK('manager','agent','mechanic','receptionist'),
          salary NUMERIC, hire_date DATE, email, phone, created_at)

vehicle_categories(id, name UNIQUE, description, daily_rate_min NUMERIC, daily_rate_max NUMERIC)

vehicles(id, category_id FK→vehicle_categories, location_id FK→locations,
         make, model, year, license_plate UNIQUE, color, daily_rate NUMERIC,
         mileage INT, status CHECK('available','rented','maintenance','retired'), created_at)

clients(id, first_name, last_name, email UNIQUE, phone,
        drivers_license, date_of_birth DATE, registration_date DATE, created_at)

reservations(id, client_id FK→clients, vehicle_id FK→vehicles,
             pickup_location FK→locations, return_location FK→locations,
             pickup_date DATE, return_date DATE,
             status CHECK('confirmed','active','completed','cancelled','no_show'),
             total_cost NUMERIC, created_at)

payments(id, reservation_id FK→reservations, amount NUMERIC,
         payment_method CHECK('credit_card','debit_card','cash'),
         payment_date DATE,
         status CHECK('completed','pending','refunded','failed'), created_at)

maintenance_records(id, vehicle_id FK→vehicles,
                    maintenance_type CHECK('oil_change','tire_rotation','brake_service',
                      'general_inspection','engine_repair','body_repair',
                      'transmission','ac_service','battery_replacement'),
                    description TEXT, cost NUMERIC, maintenance_date DATE,
                    mileage_at_service INT, completed BOOL, created_at)

reviews(id, reservation_id FK→reservations, rating INT CHECK(1-5),
        comment TEXT, review_date DATE, created_at)
""".strip()

SQL_SYSTEM_PROMPT = f"""You are a SQL expert for a car rental company database.

{SCHEMA}

When given a user question, decide if it requires querying the database.

Respond ONLY with a valid JSON object — no markdown, no explanation, no extra text.

If the question needs data:
{{"sql": "SELECT ...", "reasoning": "one-line explanation", "chart": {{"type": "bar|line|pie|area|none", "title": "Chart title", "x": "column_name_for_x_axis", "y": "column_name_for_y_axis"}}}}

If the question does NOT need data (e.g. greetings, general knowledge):
{{"sql": null, "reasoning": "no data needed", "chart": null}}

Chart rules:
- Set chart.type to "none" when the result is a single scalar, a list of text-only rows, or not meaningful as a visualization.
- Use "bar" for comparisons across categories.
- Use "line" for time-series or trends.
- Use "pie" for proportions (fewer than ~8 slices).
- Use "area" for cumulative or stacked time-series.
- x must be the column name used for labels/categories; y must be the column with numeric values.
- If multiple numeric columns exist, pick the most relevant one for y.
- The user may explicitly ask for a chart/report/export — always honour that request.

SQL rules:
- Use only the tables and columns listed above.
- Always use table aliases for clarity in JOINs.
- Limit results to 100 rows unless the user asks for more.
- Never use DELETE, UPDATE, INSERT, DROP, or any DDL/DML except SELECT.
"""

ANSWER_SYSTEM_PROMPT = """You are a helpful assistant for a car rental company.
You have access to live data from the company database.
Answer the user's question naturally and concisely based on the data provided.
If the data is a table, summarise the key findings rather than listing every row unless asked.
Format numbers clearly (e.g. currency with 2 decimal places).
Do NOT include, repeat, or show the SQL query in your response — only present the results in plain language.
"""


# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------
class MessageIn(BaseModel):
    role: str  # "user" or "assistant"
    content: str


class ChatRequest(BaseModel):
    messages: list[MessageIn]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _configure_gemini():
    if not GEMINI_API_KEY:
        raise HTTPException(status_code=500, detail="GEMINI_API_KEY not configured")
    genai.configure(api_key=GEMINI_API_KEY)
    return genai.GenerativeModel(GEMINI_MODEL)


def _build_history(messages: list[MessageIn]) -> tuple[list[dict], str]:
    """Split messages into Gemini history + last user message."""
    history = []
    for msg in messages[:-1]:
        role = "model" if msg.role == "assistant" else "user"
        history.append({"role": role, "parts": [msg.content]})
    return history, messages[-1].content


BLOCKED_KEYWORDS = {
    "INSERT", "UPDATE", "DELETE", "DROP", "CREATE", "ALTER",
    "TRUNCATE", "GRANT", "REVOKE", "EXECUTE", "CALL", "MERGE",
    "REPLACE", "UPSERT", "COPY",
}


def validate_sql(sql: str) -> str | None:
    """
    Returns None if the SQL is safe (read-only SELECT).
    Returns an error string describing the problem otherwise.
    """
    statements = sqlparse.parse(sql.strip())

    if len(statements) == 0:
        return "Empty query."

    if len(statements) > 1:
        return "Multiple statements are not allowed."

    stmt: Statement = statements[0]

    # Check the top-level statement type
    stmt_type = stmt.get_type()
    if stmt_type != "SELECT":
        return f"Statement type '{stmt_type}' is not allowed — only SELECT is permitted."

    # Walk every token and block any dangerous keywords regardless of position
    # (catches things like SELECT ... INTO, subquery abuse, etc.)
    for token in stmt.flatten():
        upper = token.normalized.upper()
        if token.ttype in (Keyword, DDL, DML) and upper in BLOCKED_KEYWORDS:
            return f"Keyword '{upper}' is not allowed."

    return None


async def _generate_sql(model, history: list[dict], last_message: str) -> dict:
    """Turn 1: ask the model for SQL. Returns {sql, reasoning}."""
    # Build a condensed history summary for context (role + first 200 chars)
    history_text = ""
    if history:
        lines = []
        for msg in history[-6:]:  # last 3 exchanges
            role = "Assistant" if msg["role"] == "model" else "User"
            text = msg["parts"][0][:200].replace("\n", " ")
            lines.append(f"{role}: {text}")
        history_text = "\n".join(lines) + "\n\n"

    chat = model.start_chat(history=[])
    prompt = f"{SQL_SYSTEM_PROMPT}\n\n{history_text}User question: {last_message}"
    response = chat.send_message(prompt)
    raw = response.text.strip()

    # Strip markdown code fences if the model adds them despite instructions
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
        raw = raw.strip()

    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        logger.warning("SQL generation returned non-JSON: %s", raw)
        return {"sql": None, "reasoning": "parse error"}


async def _execute_sql(sql: str) -> dict:
    """Call the test_app_backend SQL runner."""
    async with httpx.AsyncClient(timeout=15.0) as client:
        resp = await client.post(
            f"{SQL_BACKEND_URL}/api/sql",
            json={"query": sql},
        )
        resp.raise_for_status()
        return resp.json()


def _format_results(sql_result: dict) -> str:
    """Render SQL results as a compact markdown table for the answer prompt."""
    columns = sql_result.get("columns", [])
    rows = sql_result.get("rows", [])
    row_count = sql_result.get("row_count", 0)
    duration_ms = sql_result.get("duration_ms", 0)

    if not columns:
        return f"Query returned no rows. ({duration_ms:.1f} ms)"

    header = " | ".join(columns)
    separator = " | ".join(["---"] * len(columns))
    body = "\n".join(" | ".join(str(cell) for cell in row) for row in rows)

    return (
        f"Query returned {row_count} row(s) in {duration_ms:.1f} ms:\n\n"
        f"| {header} |\n| {separator} |\n"
        + "\n".join(f"| {' | '.join(str(c) for c in row)} |" for row in rows)
    )


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------
@app.get("/health")
async def health():
    return {"status": "ok", "model": GEMINI_MODEL}


@app.post("/chat")
async def chat(req: ChatRequest):
    if not req.messages:
        raise HTTPException(status_code=400, detail="No messages provided")

    model = _configure_gemini()
    history, last_message = _build_history(req.messages)

    # -- Turn 1: SQL generation --
    sql_result_text = None
    sql_query = None
    chart_config = None
    raw_result = None
    try:
        sql_data = await _generate_sql(model, history, last_message)
        sql_query = sql_data.get("sql")
        chart_raw = sql_data.get("chart")
        if isinstance(chart_raw, dict) and chart_raw.get("type") not in (None, "none"):
            chart_config = chart_raw
        logger.info("SQL decision — query: %s | reasoning: %s | chart: %s", sql_query, sql_data.get("reasoning"), chart_config)
    except Exception as e:
        logger.error("SQL generation failed: %s", e)

    # -- Turn 2 (optional): validate then execute SQL --
    sql_meta: dict = {"sql": sql_query, "row_count": None, "duration_ms": None, "blocked": None}
    if sql_query:
        validation_error = validate_sql(sql_query)
        if validation_error:
            logger.warning("SQL blocked by validator: %s | query: %s", validation_error, sql_query)
            sql_result_text = f"SQL blocked: {validation_error}"
            sql_meta["blocked"] = validation_error
        else:
            try:
                raw_result = await _execute_sql(sql_query)
                sql_result_text = _format_results(raw_result)
                sql_meta["row_count"] = raw_result.get("row_count")
                sql_meta["duration_ms"] = raw_result.get("duration_ms")
                logger.info("SQL executed successfully — %d rows", raw_result.get("row_count", 0))
            except Exception as e:
                logger.error("SQL execution failed: %s", e)
                sql_result_text = f"SQL execution error: {e}"
                sql_meta["blocked"] = str(e)

    # -- Turn 3: stream the final answer --
    if sql_result_text:
        answer_context = (
            f"The user asked: {last_message}\n\n"
            f"Database results:\n{sql_result_text}"
        )
    else:
        answer_context = last_message

    answer_prompt = f"{ANSWER_SYSTEM_PROMPT}\n\n{answer_context}"

    answer_chat = model.start_chat(history=history)

    async def stream_response():
        try:
            yield f"data: [META]{json.dumps(sql_meta)}\n\n"
            if raw_result and raw_result.get("columns"):
                data_payload = {
                    "columns": raw_result["columns"],
                    "rows": raw_result["rows"],
                    "chart": chart_config,
                }
                yield f"data: [DATA]{json.dumps(data_payload)}\n\n"
            response = answer_chat.send_message(answer_prompt, stream=True)
            for chunk in response:
                if chunk.text:
                    # JSON-encode the chunk so multi-line text stays on one SSE line
                    yield f"data: {json.dumps(chunk.text)}\n\n"
            yield "data: [DONE]\n\n"
        except Exception as e:
            logger.error("Gemini streaming error: %s", e)
            yield f"data: [ERROR] {str(e)}\n\n"

    return StreamingResponse(
        stream_response(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


# ---------------------------------------------------------------------------
# Excel report generation
# ---------------------------------------------------------------------------
class ChartConfig(BaseModel):
    type: str  # bar, line, pie, area
    title: str
    x: str
    y: str


class ReportRequest(BaseModel):
    columns: list[str]
    rows: list[list]
    chart: Optional[ChartConfig] = None
    title: Optional[str] = "Report"


CHART_BUILDERS = {
    "bar": BarChart,
    "line": LineChart,
    "pie": PieChart,
    "area": AreaChart,
}


def _build_excel(req: ReportRequest) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(req.columns)
    for row in req.rows:
        ws.append(row)

    for col_idx in range(1, len(req.columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, len(req.rows) + 2)
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    if req.chart and req.chart.type in CHART_BUILDERS:
        x_idx = None
        y_idx = None
        for i, col in enumerate(req.columns):
            if col == req.chart.x:
                x_idx = i + 1
            if col == req.chart.y:
                y_idx = i + 1

        if x_idx and y_idx:
            chart_cls = CHART_BUILDERS[req.chart.type]
            chart = chart_cls()
            chart.title = req.chart.title or req.title
            chart.width = 20
            chart.height = 12

            num_rows = len(req.rows)

            if req.chart.type == "pie":
                data_ref = Reference(ws, min_col=y_idx, min_row=1, max_row=num_rows + 1)
                cat_ref = Reference(ws, min_col=x_idx, min_row=2, max_row=num_rows + 1)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cat_ref)
            else:
                data_ref = Reference(ws, min_col=y_idx, min_row=1, max_row=num_rows + 1)
                cat_ref = Reference(ws, min_col=x_idx, min_row=2, max_row=num_rows + 1)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cat_ref)
                chart.x_axis.title = req.chart.x
                chart.y_axis.title = req.chart.y

            chart_ws = wb.create_sheet("Chart")
            chart_ws.add_chart(chart, "A1")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@app.post("/report")
async def generate_report(req: ReportRequest):
    try:
        excel_bytes = _build_excel(req)
    except Exception as e:
        logger.error("Excel generation failed: %s", e)
        raise HTTPException(status_code=500, detail=str(e))

    filename = f"{req.title or 'report'}.xlsx"
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
